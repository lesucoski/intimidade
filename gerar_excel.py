from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()

# Cores
GOLD = "FFD4AF37"
WINE = "FF900C3F"
BLACK = "FF0A0A0A"
DARK = "FF141414"
WHITE = "FFEEEEEE"
GREEN = "FF4ADE80"
RED = "FFF87171"

def fill(hex): return PatternFill("solid", fgColor=hex)
def font(bold=False, color="FFEEEEEE", size=11): return Font(bold=bold, color=color, name="Calibri", size=size)
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left(): return Alignment(horizontal="left", vertical="center")
def right(): return Alignment(horizontal="right", vertical="center")
def border():
    s = Side(style="thin", color="FF333333")
    return Border(left=s, right=s, top=s, bottom=s)
def money(ws, row, col, formula):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = 'R$ #,##0.00'
    return c
def pct(ws, row, col, formula):
    c = ws.cell(row=row, column=col, value=formula)
    c.number_format = '0.0"%"'
    return c

def header(ws, text):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    c = ws['A1']
    c.value = text
    c.fill = fill(BLACK)
    c.font = Font(bold=True, color=GOLD, name="Calibri", size=16)
    c.alignment = center()
    ws.row_dimensions[1].height = 36

def section(ws, row, col, text, span=4):
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+span-1)
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(DARK)
    c.font = Font(bold=True, color=GOLD, name="Calibri", size=10)
    c.alignment = left()
    ws.row_dimensions[row].height = 20

def row_label(ws, row, col, label, val, fmt='money', gold=False, bold=False):
    lc = ws.cell(row=row, column=col, value=label)
    lc.fill = fill("FF1A1A1A")
    lc.font = Font(color="FF999999", name="Calibri", size=10, bold=bold)
    lc.alignment = left()
    vc = ws.cell(row=row, column=col+1, value=val)
    vc.fill = fill("FF1A1A1A")
    vc.font = Font(color=GOLD if gold else WHITE, name="Calibri", size=10, bold=bold)
    vc.alignment = right()
    if fmt == 'money': vc.number_format = 'R$ #,##0.00'
    elif fmt == 'pct': vc.number_format = '0.0"%"'
    return vc

# ============================================================
# ABA 1 — PAINEL GERAL
# ============================================================
ws1 = wb.active
ws1.title = "📊 Painel Geral"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 4
ws1.column_dimensions['D'].width = 32
ws1.column_dimensions['E'].width = 18

header(ws1, "INTIMIDADE — Estudo de Viabilidade Financeira")

# --- ENTRADAS EDITÁVEIS ---
section(ws1, 3, 1, "⚙️  ENTRADAS — EDITE AQUI", 2)
labels_in = [
    ("Preço de Venda (R$)", 249),
    ("Unidades Vendidas / Mês", 20),
    ("Taxa Plataforma (%)", 16),
    ("Taxa Pagamento (%)", 4.99),
    ("Frete Pago pelo Vendedor (R$)", 22),
    ("Custo Fixo Mensal (R$)", 800),
    ("Marketing Mensal (R$)", 400),
    ("Alíquota Imposto (%)", 5.07),
]
for i, (lbl, val) in enumerate(labels_in, start=4):
    c_l = ws1.cell(row=i, column=1, value=lbl)
    c_l.fill = fill("FF1A1A1A"); c_l.font = font(size=10); c_l.alignment = left()
    c_v = ws1.cell(row=i, column=2, value=val)
    c_v.fill = fill("FF222222"); c_v.font = Font(bold=True, color=GOLD, name="Calibri", size=11)
    c_v.alignment = right()
    if lbl.endswith("(R$)"): c_v.number_format = 'R$ #,##0.00'
    elif lbl.endswith("(%)"): c_v.number_format = '0.00"%"'

# --- CUSTO DO KIT ---
section(ws1, 3, 4, "📦  CUSTO DO KIT", 2)
kit_items = [
    ("Óleo de Massagem", 35),
    ("Vela Aromática", 28),
    ("Máscara Tapa-olhos", 20),
    ("Embalagem Premium", 18),
    ("Fita / Laço / Adesivo", 5),
    ("Mão de Obra (Montagem)", 10),
    ("Material de Proteção", 4),
]
kit_rows = []
for i, (nome, custo) in enumerate(kit_items, start=4):
    c_l = ws1.cell(row=i, column=4, value=nome)
    c_l.fill = fill("FF1A1A1A"); c_l.font = font(size=10); c_l.alignment = left()
    c_v = ws1.cell(row=i, column=5, value=custo)
    c_v.fill = fill("FF222222"); c_v.font = Font(bold=True, color=GOLD, name="Calibri", size=11)
    c_v.alignment = right(); c_v.number_format = 'R$ #,##0.00'
    kit_rows.append(f"E{i+0}")

kit_total_row = 4 + len(kit_items)
c_tl = ws1.cell(row=kit_total_row, column=4, value="TOTAL CUSTO / KIT")
c_tl.fill = fill(DARK); c_tl.font = Font(bold=True, color=GOLD, name="Calibri", size=10); c_tl.alignment = left()
c_tv = ws1.cell(row=kit_total_row, column=5, value=f"=SUM(E4:E{kit_total_row-1})")
c_tv.fill = fill(DARK); c_tv.font = Font(bold=True, color=GOLD, name="Calibri", size=11)
c_tv.alignment = right(); c_tv.number_format = 'R$ #,##0.00'

# --- RESULTADOS ---
res_start = max(kit_total_row, 4+len(labels_in)) + 2
section(ws1, res_start, 1, "📈  RESULTADOS CALCULADOS", 5)
r = res_start + 1

# Fórmulas referenciando entradas
def v(row): return f"B{row+3}"  # offset pois entradas começam na linha 4

results = [
    ("Receita Bruta", f"=B4*B5", 'money', False),
    ("(-) Custo Variável Total", f"=E{kit_total_row}*B5", 'money', False),
    ("(-) Taxa Plataforma", f"=B4*(B6/100)*B5", 'money', False),
    ("(-) Taxa Pagamento", f"=B4*(B7/100)*B5", 'money', False),
    ("(-) Frete Total", f"=B8*B5", 'money', False),
    ("(-) Custos Fixos + Marketing", f"=B9+B10", 'money', False),
    ("(-) Impostos", f"=B4*B5*(B11/100)", 'money', False),
    ("", None, None, False),
    ("LUCRO LÍQUIDO / MÊS", f"=B{r}-B{r+1}-B{r+2}-B{r+3}-B{r+4}-B{r+5}-B{r+6}", 'money', True),
    ("MARGEM LÍQUIDA (%)", f"=IF(B{r}>0,(B{r+8}/B{r})*100,0)", 'pct', True),
    ("LUCRO POR KIT", f"=B{r+8}/B5", 'money', True),
]

formula_rows = {}
for i, (lbl, formula, fmt, bold) in enumerate(results):
    if not lbl:
        r += 1
        continue
    is_gold = bold
    c_l = ws1.cell(row=r, column=1, value=lbl)
    c_l.fill = fill(DARK if bold else "FF1A1A1A")
    c_l.font = Font(bold=bold, color=GOLD if is_gold else "FF999999", name="Calibri", size=10)
    c_l.alignment = left()
    c_v = ws1.cell(row=r, column=2, value=formula)
    c_v.fill = fill(DARK if bold else "FF1A1A1A")
    c_v.font = Font(bold=bold, color=GOLD if is_gold else WHITE, name="Calibri", size=11)
    c_v.alignment = right()
    if fmt == 'money': c_v.number_format = 'R$ #,##0.00'
    elif fmt == 'pct': c_v.number_format = '0.0"%"'
    formula_rows[lbl] = r
    r += 1

# ============================================================
# ABA 2 — CENÁRIOS
# ============================================================
ws2 = wb.create_sheet("🎯 Cenários")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions['A'].width = 16
for col in ['B','C','D','E','F']: ws2.column_dimensions[col].width = 18
header(ws2, "INTIMIDADE — Cenários de Vendas")
section(ws2, 3, 1, "Cenário", 1)
headers_c = ["Unidades/Mês","Receita Bruta","Lucro Mensal","Margem (%)","Viabilidade"]
for i, h in enumerate(headers_c, 2):
    c = ws2.cell(row=3, column=i, value=h)
    c.fill = fill(DARK); c.font = Font(bold=True, color=GOLD, name="Calibri", size=10); c.alignment = center()

cenarios = [("Pessimista", 0.5), ("Realista", 1.0), ("Otimista", 1.5), ("Agressivo", 2.5)]
kit_cost_ref = f"'📊 Painel Geral'!E{kit_total_row}"
preco_ref = "'📊 Painel Geral'!B4"
fixo_ref = "'📊 Painel Geral'!B9+'📊 Painel Geral'!B10"
tax_ref = f"({preco_ref}*('📊 Painel Geral'!B6+'📊 Painel Geral'!B7)/100)+'📊 Painel Geral'!B8"
imp_ref = f"{preco_ref}*('📊 Painel Geral'!B11/100)"

for i, (nome, mult) in enumerate(cenarios, 4):
    unid = f"=ROUND('📊 Painel Geral'!B5*{mult},0)"
    receita = f"={preco_ref}*B{i}"
    lucro = f"=C{i}-({kit_cost_ref}*B{i})-({tax_ref}*B{i})-({imp_ref}*B{i})-({fixo_ref})"
    margem = f"=IF(C{i}>0,D{i}/C{i}*100,0)"
    viab = f'=IF(E{i}>25,"✅ Viável",IF(E{i}>10,"⚠️ Marginal","❌ Inviável"))'
    row_color = "FF1A1A1A" if i % 2 == 0 else "FF141414"
    vals = [nome, unid, receita, lucro, margem, viab]
    fmts = [None, '#,##0', 'R$ #,##0.00', 'R$ #,##0.00', '0.0"%"', None]
    for j, (val, fmt) in enumerate(zip(vals, fmts), 1):
        c = ws2.cell(row=i, column=j, value=val)
        c.fill = fill(row_color)
        c.font = Font(color=GOLD if j == 1 else WHITE, name="Calibri", size=10, bold=(j==1))
        c.alignment = center()
        if fmt: c.number_format = fmt

# ============================================================
# ABA 3 — PAYBACK E ROI
# ============================================================
ws3 = wb.create_sheet("📈 Payback & ROI")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions['A'].width = 34; ws3.column_dimensions['B'].width = 20
header(ws3, "INTIMIDADE — Investimento Inicial & Payback")
section(ws3, 3, 1, "💰  INVESTIMENTO INICIAL", 2)
inv_items = [
    ("Estoque Inicial (30 kits)", f"=30*'📊 Painel Geral'!E{kit_total_row}"),
    ("Domínio + Hospedagem", 200),
    ("Material Visual / Fotos", 500),
    ("Capital de Giro Extra", 300),
    ("Outros", 200),
]
for i, (lbl, val) in enumerate(inv_items, 4):
    cl = ws3.cell(row=i, column=1, value=lbl); cl.fill = fill("FF1A1A1A"); cl.font = font(size=10); cl.alignment = left()
    cv = ws3.cell(row=i, column=2, value=val); cv.fill = fill("FF222222"); cv.font = Font(bold=True, color=GOLD, name="Calibri", size=11)
    cv.alignment = right(); cv.number_format = 'R$ #,##0.00'

inv_total_row = 4 + len(inv_items)
ctl = ws3.cell(row=inv_total_row, column=1, value="TOTAL INVESTIMENTO"); ctl.fill = fill(DARK); ctl.font = Font(bold=True, color=GOLD, name="Calibri", size=11); ctl.alignment = left()
ctv = ws3.cell(row=inv_total_row, column=2, value=f"=SUM(B4:B{inv_total_row-1})"); ctv.fill = fill(DARK); ctv.font = Font(bold=True, color=GOLD, name="Calibri", size=12); ctv.alignment = right(); ctv.number_format = 'R$ #,##0.00'

section(ws3, inv_total_row+2, 1, "📊  ANÁLISE DE RETORNO", 2)
# Referências ao Painel Geral (lucro mensal está na linha r-1 do painel)
lucro_ref = f"'📊 Painel Geral'!B{formula_rows.get('LUCRO LÍQUIDO / MÊS', 20)}"
margem_ref = f"'📊 Painel Geral'!B{formula_rows.get('MARGEM LÍQUIDA (%)', 21)}"
lucro_kit_ref = f"'📊 Painel Geral'!B{formula_rows.get('LUCRO POR KIT', 22)}"

retorno = [
    ("Lucro Mensal Esperado", lucro_ref, 'money'),
    ("Margem Líquida", margem_ref, 'pct'),
    ("Lucro por Kit", lucro_kit_ref, 'money'),
    ("Break-Even (unid./mês)", f"=('📊 Painel Geral'!B9+'📊 Painel Geral'!B10)/MAX({lucro_kit_ref},0.01)", '#,##0.0'),
    ("Payback (meses)", f"=IF({lucro_ref}>0,B{inv_total_row}/{lucro_ref},999)", '0.0'),
    ("ROI Anual (%)", f"=IF(B{inv_total_row}>0,({lucro_ref}*12/B{inv_total_row})*100,0)", '0.0"%"'),
    ("Lucro Anual Estimado", f"={lucro_ref}*12", 'money'),
]
for i, (lbl, val, fmt) in enumerate(retorno, inv_total_row+3):
    cl = ws3.cell(row=i, column=1, value=lbl); cl.fill = fill("FF1A1A1A"); cl.font = font(size=10); cl.alignment = left()
    cv = ws3.cell(row=i, column=2, value=val); cv.fill = fill("FF1A1A1A"); cv.font = Font(bold=True, color=GOLD, name="Calibri", size=11); cv.alignment = right()
    cv.number_format = fmt if fmt != 'money' else 'R$ #,##0.00'
    if fmt == 'pct': cv.number_format = '0.0"%"'

# ============================================================
# ABA 4 — TABELA DE KITS
# ============================================================
ws4 = wb.create_sheet("📦 Kits & Produtos")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions['A'].width = 28; ws4.column_dimensions['B'].width = 22; ws4.column_dimensions['C'].width = 16; ws4.column_dimensions['D'].width = 16
header(ws4, "INTIMIDADE — Catálogo de Produtos e Kits")
section(ws4, 3, 1, "Produto", 1)
for j, h in enumerate(["Fornecedor/Distribuidor","Custo Unit. (R$)","Incluso no Kit"], 2):
    c = ws4.cell(row=3, column=j, value=h); c.fill = fill(DARK); c.font = Font(bold=True, color=GOLD, name="Calibri", size=10); c.alignment = center()

produtos = [
    ("Óleo de Massagem 100ml", "Fornecedor A - Atacado", 35, "✅"),
    ("Vela Aromática Soja", "Fornecedor A - Atacado", 28, "✅"),
    ("Máscara Tapa-olhos Cetim", "Fornecedor B - Importado", 20, "✅"),
    ("Embalagem Premium Black", "Gráfica Local", 18, "✅"),
    ("Fita de Cetim + Adesivo", "Papelaria Atacado", 5, "✅"),
    ("Pétala de Rosa Seca", "Flora Atacado", 8, "⭕ Opcional"),
    ("Cartão Personalizado", "Gráfica Local", 4, "⭕ Opcional"),
    ("Sachê Perfumado", "Fornecedor B - Importado", 12, "⭕ Opcional"),
]
for i, (nome, dist, custo, incl) in enumerate(produtos, 4):
    row_c = "FF1A1A1A" if i%2==0 else "FF141414"
    for j, (val, fmt) in enumerate([(nome,None),(dist,None),(custo,'R$ #,##0.00'),(incl,None)], 1):
        c = ws4.cell(row=i, column=j, value=val); c.fill = fill(row_c)
        c.font = Font(color=GOLD if j==3 else WHITE, name="Calibri", size=10); c.alignment = center() if j>1 else left()
        if fmt: c.number_format = fmt

# ============================================================
# ABA 5 — IMPOSTOS E TAXAS
# ============================================================
ws5 = wb.create_sheet("🏛️ Impostos & Taxas")
ws5.sheet_view.showGridLines = False
ws5.column_dimensions['A'].width = 30; ws5.column_dimensions['B'].width = 14; ws5.column_dimensions['C'].width = 35
header(ws5, "INTIMIDADE — Guia de Impostos e Taxas de Marketplace")
section(ws5, 3, 1, "Regime Tributário", 3)
impostos = [
    ("MEI", "~R$ 75/mês fixo", "CNPJ simples. Limite R$81k/ano. Ideal para início."),
    ("Simples Nacional - Anexo I (Comércio)", "4% a 11,2%", "Sobre faturamento. Recomendado a partir de R$7k/mês."),
    ("Lucro Presumido", "~8% + 3% = ~11%", "Para faturamento acima de R$360k/ano."),
]
for i, (regime, aliq, obs) in enumerate(impostos, 4):
    for j, val in enumerate([regime, aliq, obs], 1):
        c = ws5.cell(row=i, column=j, value=val); c.fill = fill("FF1A1A1A" if i%2==0 else "FF141414")
        c.font = Font(color=GOLD if j==1 else ("FF999999" if j==3 else WHITE), name="Calibri", size=10)
        c.alignment = left(); c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

section(ws5, 9, 1, "Taxas por Plataforma", 3)
plataformas = [
    ("Mercado Livre - Clássico","10-11% + 4,99% pgto","Visibilidade menor. Sem frete grátis obrigatório."),
    ("Mercado Livre - Premium","16-19% + 4,99% pgto","Topo de busca. Frete grátis obrigatório acima R$79."),
    ("Shopee","12-16%","Frete subsidiado em campanhas. Sem taxa de pgto extra."),
    ("WhatsApp / Site Próprio","0% + 3,5% gateway","Melhor margem. Você gerencia o atendimento."),
    ("Amazon BR","~15%","Marketplace premium. Exige cadastro como profissional."),
]
for i, (plat, taxa, obs) in enumerate(plataformas, 10):
    for j, val in enumerate([plat, taxa, obs], 1):
        c = ws5.cell(row=i, column=j, value=val); c.fill = fill("FF1A1A1A" if i%2==0 else "FF141414")
        c.font = Font(color=GOLD if j==1 else ("FF999999" if j==3 else WHITE), name="Calibri", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws5.row_dimensions[i].height = 32

# Ajustar altura das linhas
for ws in [ws1, ws2, ws3, ws4, ws5]:
    for row in ws.iter_rows():
        for cell in row:
            if not cell.font:
                cell.font = font()

output = r"C:\Users\LeandroCarvalho\.gemini\antigravity\scratch\intimidade\Intimidade_Viabilidade.xlsx"
wb.save(output)
print("Excel gerado com sucesso em:")
print(output)
